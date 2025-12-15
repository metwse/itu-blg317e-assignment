"""Load economies table (names, codes, regions)."""

from src.entities import Economy, Region
from src.state import State

from typing import cast
from openpyxl import load_workbook
import httpx
import io


URL = "https://ddh-openapi.worldbank.org/resources/DR0095333/download"


async def load(state: State):
    async with httpx.AsyncClient() as client:
        response = await client.get(URL, follow_redirects=True)
        response.raise_for_status()

    wb = load_workbook(filename=io.BytesIO(response.content),
                       read_only=True, data_only=True)
    sheet = wb.active

    if sheet is None:
        raise ValueError('Could not fetch sheet')

    rows = sheet.iter_rows(values_only=True)

    headers = next(rows)
    header_map = {h: i for i, h in enumerate(headers) if h}

    code_idx = header_map.get("Code")
    name_idx = header_map.get("Economy")
    region_idx = header_map.get("Region")

    if code_idx is None or name_idx is None or region_idx is None:
        raise ValueError('Could not parse xlsx')

    for row in rows:
        if not row:
            continue

        code = row[code_idx]
        name = row[name_idx]
        region = row[region_idx]

        if code is not None and name is not None:
            await state.economy_service.create(
                Economy(code=cast(str, code),
                        name=cast(str, name),
                        region=cast(Region | None, region)))
